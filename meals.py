"""
meals.py — Recipe and ingredient data for the Family Meal Planner bot.

Reads directly from Ingredient.xlsx in the same folder.
Edit the spreadsheet to update recipes — no code changes needed.

Excel layout:
  Row 4:  Ingredient header (Meat/Protein | Vegetables | Side | Condiment | Fruits | Carbs | Soup)
  Row 6+: Ingredient master list (columns B–H)
  Row 19: Recipe header row  ← skipped
  Row 20: "Protein" section header
  Row 21+: Recipe rows (B=name, C-F=ingredients)
  Row 30: "Vegetable" section header
  Row 42: "Soup" section header
"""

import os
import random
from openpyxl import load_workbook

SPREADSHEET = os.path.join(os.path.dirname(__file__), "Ingredient.xlsx")

# Rows to explicitly skip (headers, blank separators)
SKIP_ROWS = {19}  # "Recipe | Ingredient 1 | Ingredient 2..." header row

# Values that should never be treated as recipe names or ingredients
SKIP_VALUES = {"none", "ingredient 1", "ingredient 2", "ingredient 3",
               "ingredient 4", "recipe", "ingredients", "meat / protein",
               "protein", "vegetable", "soup", "side", "condiment",
               "fruits", "carbs", "carbs "}

# Known fruits — only pick from column F (index 5) if value is in this list
# Tomato, Capsicum etc in column F of ingredient rows are NOT fruits
KNOWN_FRUITS = {
    "banana", "dragonfruit", "blueberries", "apple", "pear",
    "raspberry", "strawberry", "avocado banana", "watermelon",
    "mango", "papaya", "orange", "kiwi", "grapes",
}

# Weekend-only fruits
WEEKEND_FRUITS = {"Raspberry", "Strawberry", "Avocado banana", "Blueberries"}

# Soup name suffix — all soups get " soup" appended if not already present
def _soup_name(raw: str) -> str:
    r = raw.strip()
    if r.lower().endswith("soup") or r.lower().endswith("stock") or r.lower().endswith("tang"):
        return r
    return r + " soup"


def load_recipes() -> dict:
    """
    Returns:
    {
        "protein": [{"name": ..., "ing": [...]}, ...],
        "veg":     [...],
        "soup":    [...],   # names all end with "soup" / "stock" / "tang"
        "fruit":   [...],
        "carbs":   [...],
    }
    """
    wb = load_workbook(SPREADSHEET, read_only=True, data_only=True)
    ws = wb.active

    recipes  = {"protein": [], "veg": [], "soup": [], "fruit": [], "carbs": []}
    current_section = None

    SECTION_MAP = {
        "protein":  "protein",
        "vegetable": "veg",
        "soup":     "soup",
    }

    fruits_seen = []   # ordered, deduped
    carbs_seen  = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if row_idx in SKIP_ROWS:
            continue

        b = str(row[1]).strip() if row[1] else ""
        b_low = b.lower()

        # ── Collect fruits from column F (index 5) ──
        # Only accept values that are actually known fruits
        f_raw = str(row[5]).strip() if row[5] else ""
        f_low = f_raw.lower().strip()
        if f_low and f_low in KNOWN_FRUITS and f_raw not in fruits_seen:
            fruits_seen.append(f_raw.strip())

        # ── Collect carbs from column G (index 6) ──
        c_raw = str(row[6]).strip() if row[6] else ""
        c_low = c_raw.lower().strip()
        if c_low and c_low not in SKIP_VALUES and c_raw not in carbs_seen:
            carbs_seen.append(c_raw.strip())

        # ── Section detection ──
        if b_low in SECTION_MAP:
            current_section = SECTION_MAP[b_low]
            continue

        # ── Skip header/blank rows ──
        if not b or b_low in SKIP_VALUES:
            continue

        # ── Skip broken formula cells ──
        if b.startswith("="):
            b = "Spinach"

        # ── Add recipe to current section ──
        if current_section:
            ing = []
            for i in range(2, 6):  # columns C–F = indices 2–5
                val = str(row[i]).strip() if row[i] else ""
                if val and val.lower() not in SKIP_VALUES:
                    ing.append(val)

            # For soups: rename to include "soup" / "stock" / "tang"
            if current_section == "soup":
                b = _soup_name(b)

            existing_names = [r["name"] for r in recipes[current_section]]
            if b not in existing_names:
                recipes[current_section].append({"name": b, "ing": ing})

    # ── Build fruit list ──
    for f in fruits_seen:
        recipes["fruit"].append({"name": f, "ing": [f]})

    # Ensure weekend fruits always present
    existing_fruits = {r["name"] for r in recipes["fruit"]}
    for wf in sorted(WEEKEND_FRUITS):
        if wf not in existing_fruits:
            recipes["fruit"].append({"name": wf, "ing": [wf]})

    # ── Build carbs list ──
    for c in carbs_seen:
        recipes["carbs"].append({"name": c, "ing": [c]})

    # Ensure Pasta with prawn always present
    carb_names = {r["name"] for r in recipes["carbs"]}
    if "Pasta with prawn" not in carb_names:
        recipes["carbs"].append({"name": "Pasta with prawn", "ing": ["Pasta", "Prawn"]})

    # Also add Chicken Breast as a protein if not present (used by Monday fixed rule)
    protein_names = {r["name"] for r in recipes["protein"]}
    if "Chicken Breast" not in protein_names:
        recipes["protein"].append({"name": "Chicken Breast", "ing": ["Chicken Breast", "Ginger"]})
    if "Salmon egg" not in protein_names:
        recipes["protein"].append({"name": "Salmon egg", "ing": ["Salmon", "Eggs"]})

    wb.close()
    return recipes


# ─────────────────────────────────────────────────────────────────────────────

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
IS_WEEKEND = {"Saturday": True, "Sunday": True}


def generate_week(recipes: dict) -> dict:
    """
    Weekly meal plan:

    DINNER (fixed rotation):
      Monday    — Dun Ji Tang soup + Chicken Breast + veg
      Tuesday   — Salmon egg + veg + pork rib soup
      Wednesday — Pork/chicken protein + veg + egg dish
      Thursday  — Pork rib soup + chicken/egg protein + veg
      Friday    — Varied protein + veg (no soup)
      Sat/Sun   — Protein + veg (no soup)

    LUNCH: random protein + veg each day
      Saturday lunch = Pasta with prawn (1 dish only, no other dish)

    SALMON RULE: max 2x per week
      - Tuesday dinner is always Salmon egg (fixed)
      - One more salmon dish allowed on Thu or Fri dinner only
      - No other salmon anywhere

    SOUP: 1x Dun Ji Tang (Mon), 2x pork rib soups (Tue + Thu)

    FRUIT:
      Mon/Tue  — weekday fruit, NOT Apple
      Wed/Thu  — different weekday fruit
      Fri      — weekday fruit, NOT Apple
      Sat/Sun  — weekend fruit only (Raspberry/Blueberries/Strawberry/Avocado banana)
      Weekend breakfast: Oats + fruit

    VEG: Broccoli 3–4x, others max 2x each
    """

    def rnd(pool, excl=[]):
        a = [r for r in pool if r["name"] not in excl]
        return random.choice(a if a else pool)

    def find(type_, name):
        for r in recipes.get(type_, []):
            if r["name"] == name:
                return {"type": type_, "name": name, "ing": r["ing"]}
        return {"type": type_, "name": name, "ing": []}

    # ── Veg rotation ──────────────────────────────────────────────────────────
    used_veg = {}

    def weighted_veg(excl=[]):
        pool = []
        for r in recipes["veg"]:
            if r["name"] not in excl:
                w = 4 if "brocoli" in r["name"].lower() else 1
                pool.extend([r] * w)
        return random.choice(pool if pool else recipes["veg"])

    def nv(excl=[]):
        over = [n for n, c in used_veg.items()
                if (c >= 4 if "brocoli" in n.lower() else c >= 2)]
        r = weighted_veg(excl=over + excl)
        used_veg[r["name"]] = used_veg.get(r["name"], 0) + 1
        return {"type": "veg", **r}

    # ── Protein rotation ──────────────────────────────────────────────────────
    used_prot = []
    salmon_count = [0]  # track salmon dishes; max 2 (Tue dinner + one more)

    def is_salmon(name):
        return "salmon" in name.lower()

    def np(excl=[], allow_salmon=False):
        pool = recipes["protein"]
        # Exclude salmon if already at max or not allowed
        salmon_excl = [] if (allow_salmon and salmon_count[0] < 2) else \
                      [r["name"] for r in pool if is_salmon(r["name"])]
        r = rnd(pool, used_prot + excl + salmon_excl)
        used_prot.append(r["name"])
        if len(used_prot) > 3: used_prot.pop(0)
        if is_salmon(r["name"]): salmon_count[0] += 1
        return {"type": "protein", **r}

    # ── Fruit rotation ────────────────────────────────────────────────────────
    WKND_SET = {"Raspberry", "Strawberry", "Avocado banana", "Blueberries"}
    weekday_f  = [f for f in recipes["fruit"] if f["name"] not in WKND_SET]
    weekend_f  = [f for f in recipes["fruit"] if f["name"] in WKND_SET]
    non_apple  = [f for f in weekday_f if f["name"].strip().lower() != "apple"]

    if not weekend_f:
        weekend_f = recipes["fruit"]
    if not non_apple:
        non_apple = weekday_f

    f_mon  = random.choice(non_apple)
    f_wed  = rnd(weekday_f, [f_mon["name"]])
    f_fri  = rnd([f for f in non_apple if f["name"] != f_mon["name"]] or non_apple, [])
    f_wknd = random.choice(weekend_f)

    fruit_by_day = {
        "Monday": f_mon, "Tuesday": f_mon,
        "Wednesday": f_wed, "Thursday": f_wed,
        "Friday": f_fri,
        "Saturday": f_wknd, "Sunday": f_wknd,
    }

    def breakfast(day):
        fruit = fruit_by_day[day]
        fd = {"type": "fruit", "name": fruit["name"], "ing": fruit["ing"]}
        if IS_WEEKEND.get(day):
            return [{"type": "carbs", "name": "Oats", "ing": ["Oats", fruit["name"]]}, fd]
        return [fd]

    # ── Soup selection ────────────────────────────────────────────────────────
    pork_soups = [r for r in recipes["soup"] if "Pork Ribs" in r["ing"]]
    random.shuffle(pork_soups)
    ps_tue = pork_soups[0] if pork_soups else None
    ps_thu = pork_soups[1] if len(pork_soups) > 1 else ps_tue

    # ── Wed dinner proteins ───────────────────────────────────────────────────
    pork_chk = [r for r in recipes["protein"]
                if any(i in r["ing"] for i in
                       ["Minced Pork", "Chicken Thighs", "Chicken drumstick"])]
    egg_prot = [r for r in recipes["protein"]
                if "Eggs" in r["ing"] or "egg" in r["name"].lower()]
    wed_pork = rnd(pork_chk, ["Chicken Breast", "Salmon egg"]) if pork_chk else rnd(recipes["protein"], [])
    # Wed egg: exclude salmon entirely (salmon is reserved for Tue dinner + one Thu/Fri slot)
    non_salmon_egg = [r for r in egg_prot if not is_salmon(r["name"])]
    wed_egg = rnd(non_salmon_egg if non_salmon_egg else egg_prot, [wed_pork["name"]])

    # ── Thu dinner protein ────────────────────────────────────────────────────
    # Thu pool excludes salmon (salmon only appears via explicit second_salmon_day logic)
    thu_pool = [r for r in recipes["protein"]
                if any(i in r["ing"] for i in
                       ["Chicken Thighs", "Chicken Breast", "Chicken drumstick", "Eggs"])
                and not is_salmon(r["name"])]
    thu_prot = rnd(thu_pool, [wed_pork["name"], wed_egg["name"]]) if thu_pool else rnd(recipes["protein"], [])

    # ── Decide if Thu or Fri dinner gets the 2nd salmon dish ─────────────────
    second_salmon_day = random.choice(["Thursday", "Friday"])

    # ── Build plan ────────────────────────────────────────────────────────────
    plan = {}
    for day in DAYS:
        plan[day] = {
            "out": {"Breakfast": False, "Lunch": False, "Dinner": False},
            "Breakfast": breakfast(day),
            "Lunch": [], "Dinner": [],
        }

    # Count Tuesday salmon egg as 1
    salmon_count[0] = 1  # pre-count Tue dinner salmon egg

    # Monday
    plan["Monday"]["Lunch"]  = [np(excl=["Chicken Breast"]), nv()]
    plan["Monday"]["Dinner"] = [find("soup", "Dun Ji Tang"),
                                find("protein", "Chicken Breast"), nv()]

    # Tuesday
    plan["Tuesday"]["Lunch"]  = [np(excl=["Salmon egg"]), nv()]
    plan["Tuesday"]["Dinner"] = [find("protein", "Salmon egg"), nv(),
                                 {"type": "soup", **ps_tue} if ps_tue else np()]

    # Wednesday
    plan["Wednesday"]["Lunch"]  = [np(excl=[wed_pork["name"], wed_egg["name"]]), nv()]
    plan["Wednesday"]["Dinner"] = [{"type": "protein", **wed_pork},
                                   nv(),
                                   {"type": "protein", **wed_egg}]

    # Thursday
    salmon_allowed_thu = (second_salmon_day == "Thursday")
    plan["Thursday"]["Lunch"]  = [np(excl=[thu_prot["name"]]), nv()]
    plan["Thursday"]["Dinner"] = [
        {"type": "soup", **ps_thu} if ps_thu else np(),
        find("protein", "Steamed Salmon") if salmon_allowed_thu
            else {"type": "protein", **thu_prot},
        nv(),
    ]
    if salmon_allowed_thu:
        salmon_count[0] += 1

    # Friday
    salmon_allowed_fri = (second_salmon_day == "Friday")
    plan["Friday"]["Lunch"]  = [np(excl=["Salmon egg"], allow_salmon=False), nv()]
    plan["Friday"]["Dinner"] = [
        find("protein", "Steamed Salmon") if salmon_allowed_fri
            else np(excl=["Salmon egg"], allow_salmon=False),
        nv(),
    ]

    # Saturday — pasta with prawn lunch only
    plan["Saturday"]["Lunch"]  = [{"type": "carbs", "name": "Pasta with prawn",
                                    "ing": ["Pasta", "Prawn"]}]
    plan["Saturday"]["Dinner"] = [np(allow_salmon=False), nv()]

    # Sunday
    plan["Sunday"]["Lunch"]  = [np(allow_salmon=False), nv()]
    plan["Sunday"]["Dinner"] = [np(allow_salmon=False), nv()]

    return plan


# ── Grocery split ─────────────────────────────────────────────────────────────

MON_DAYS = ["Monday", "Tuesday", "Wednesday"]
THU_DAYS = ["Thursday", "Friday", "Saturday", "Sunday"]


def build_grocery(plan: dict) -> dict:
    def collect(days):
        items = set()
        for day in days:
            data = plan[day]
            for meal in ["Breakfast", "Lunch", "Dinner"]:
                if data["out"].get(meal):
                    continue
                for dish in data.get(meal, []):
                    ings = dish.get("ing") or [dish["name"]]
                    for i in ings:
                        if i:
                            items.add(i.strip())
        return sorted(items)

    return {"mon": collect(MON_DAYS), "thu": collect(THU_DAYS)}
